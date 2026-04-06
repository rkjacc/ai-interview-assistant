"""
Excel Exporter Module
Generates professional Excel spreadsheet output for interview Q&A with formatting.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import io


class ExcelExporter:
    """Generate professionally formatted Excel spreadsheets for Q&A results."""

    # Color scheme for professional appearance
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TOPIC_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOPIC_FONT = Font(bold=True, size=11)
    QUESTION_FILL = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    ANSWER_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    BORDER_STYLE = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def create_qa_workbook(qa_items: List[Dict[str, str]], 
                          candidate_info: Optional[Dict] = None) -> openpyxl.Workbook:
        """
        Create a professionally formatted Excel workbook with Q&A pairs.
        
        Args:
            qa_items: List of Q&A dictionaries with keys: Topic, Question, Answer
            candidate_info: Optional dictionary with candidate information
            
        Returns:
            openpyxl Workbook object
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interview Q&A"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Topic
        ws.column_dimensions['B'].width = 50  # Question
        ws.column_dimensions['C'].width = 60  # Answer
        
        current_row = 1
        
        # Add title
        ws.merge_cells(f'A{current_row}:C{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "Technical Interview Q&A"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Add timestamp
        ws.merge_cells(f'A{current_row}:C{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        timestamp_cell.font = Font(italic=True, size=10, color="666666")
        timestamp_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        
        # Add candidate info if provided
        if candidate_info:
            current_row += 1
            info_text = "Candidate Information"
            ws.merge_cells(f'A{current_row}:C{current_row}')
            info_header = ws[f'A{current_row}']
            info_header.value = info_text
            info_header.font = Font(bold=True, size=11, color="1F4E78")
            current_row += 1
            
            for key, value in candidate_info.items():
                if value:
                    ws[f'A{current_row}'] = f"{key}:"
                    ws[f'B{current_row}'] = str(value)
                    ws[f'A{current_row}'].font = Font(bold=True, size=10)
                    current_row += 1
            
            current_row += 1
        
        # Add header row
        headers = ['Topic', 'Question', 'Answer']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = ExcelExporter.BORDER_STYLE
        
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Add Q&A rows
        for idx, qa in enumerate(qa_items, 1):
            topic = qa.get('Topic', '')
            question = qa.get('Question', '')
            answer = qa.get('Answer', '')
            
            # Topic cell
            topic_cell = ws.cell(row=current_row, column=1)
            topic_cell.value = topic
            topic_cell.fill = ExcelExporter.TOPIC_FILL
            topic_cell.font = ExcelExporter.TOPIC_FONT
            topic_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            topic_cell.border = ExcelExporter.BORDER_STYLE
            
            # Question cell
            question_cell = ws.cell(row=current_row, column=2)
            question_cell.value = question
            question_cell.fill = ExcelExporter.QUESTION_FILL
            question_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            question_cell.border = ExcelExporter.BORDER_STYLE
            
            # Answer cell
            answer_cell = ws.cell(row=current_row, column=3)
            answer_cell.value = answer
            answer_cell.fill = ExcelExporter.ANSWER_FILL
            answer_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            answer_cell.border = ExcelExporter.BORDER_STYLE
            
            # Set row height based on content
            ws.row_dimensions[current_row].height = max(30, len(str(answer)) // 50 * 15)
            
            current_row += 1
        
        # Add summary at the bottom
        current_row += 1
        ws.merge_cells(f'A{current_row}:C{current_row}')
        summary_cell = ws[f'A{current_row}']
        summary_cell.value = f"Total Questions: {len(qa_items)}"
        summary_cell.font = Font(bold=True, size=10, color="1F4E78")
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return wb

    @staticmethod
    def save_to_file(workbook: openpyxl.Workbook, file_path: str) -> str:
        """
        Save workbook to file.
        
        Args:
            workbook: openpyxl Workbook object
            file_path: Path to save Excel file
            
        Returns:
            Path to saved file
        """
        workbook.save(file_path)
        return file_path

    @staticmethod
    def save_to_bytes(workbook: openpyxl.Workbook) -> bytes:
        """
        Save workbook to bytes for streaming download.
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            Bytes content of Excel file
        """
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_filename(candidate_name: Optional[str] = None) -> str:
        """
        Generate professional filename for Excel export.
        
        Args:
            candidate_name: Optional candidate name to include in filename
            
        Returns:
            Filename string
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if candidate_name:
            candidate_name = candidate_name.replace(' ', '_').replace(',', '')
            return f"interview_qa_{candidate_name}_{timestamp}.xlsx"
        else:
            return f"interview_qa_{timestamp}.xlsx"
